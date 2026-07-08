from __future__ import annotations

import csv
from pathlib import Path

from link_checker.enums import LinkStatus
from link_checker.models import ValidationResult

_SUPPORTED_OUTPUT_EXTENSIONS = {".xlsx", ".csv"}
_OUTPUT_COLUMNS = (
    "NOME DO PARTICIPANTE",
    "NOME DA EMPRESA",
    "LINK",
    "RESULTADO",
)
_DETAILS_COLUMNS = (
    "NOME DO PARTICIPANTE",
    "NOME DA EMPRESA",
    "LINK",
    "RESULTADO",
    "STATUS_INTERNO",
    "REGRA",
    "EVIDENCIA",
    "STATUS_HTTP",
    "URL_FINAL",
    "TEMPO_RESPOSTA_MS",
    "ERRO_TECNICO",
)
_TECHNICAL_COLUMNS = (
    "participante",
    "empresa",
    "link",
    "status",
    "http_status",
    "final_url",
    "response_time_seconds",
    "rule_name",
    "evidence",
    "technical_error",
)


class ReportWriter:
    def write(
        self, results: list[ValidationResult], path: Path, *, overwrite: bool = False
    ) -> None:
        suffix = path.suffix.lower()
        if suffix not in _SUPPORTED_OUTPUT_EXTENSIONS:
            raise ValueError(f"Extensao nao suportada: {suffix}. Use .csv ou .xlsx.")

        _ensure_can_write(path, overwrite)
        path.parent.mkdir(parents=True, exist_ok=True)
        rows = [self._row(r) for r in results]

        if suffix == ".csv":
            _write_csv(path, _OUTPUT_COLUMNS, rows)
        else:
            details_rows = [_details_row(result) for result in results]
            _write_xlsx(
                path,
                [
                    _Sheet("resultado", _OUTPUT_COLUMNS, rows, result_sheet=True),
                    _Sheet("detalhes", _DETAILS_COLUMNS, details_rows),
                ],
            )

    def write_technical(
        self, results: list[ValidationResult], path: Path, *, overwrite: bool = False
    ) -> None:
        suffix = path.suffix.lower()
        if suffix not in _SUPPORTED_OUTPUT_EXTENSIONS:
            raise ValueError(f"Extensao nao suportada: {suffix}. Use .csv ou .xlsx.")

        _ensure_can_write(path, overwrite)
        path.parent.mkdir(parents=True, exist_ok=True)
        rows = [self._technical_row(r) for r in results]

        if suffix == ".csv":
            _write_csv(path, _TECHNICAL_COLUMNS, rows)
        else:
            _write_xlsx(path, [_Sheet("resultado_tecnico", _TECHNICAL_COLUMNS, rows)])

    @staticmethod
    def _row(result: ValidationResult) -> list[object]:
        return [
            result.participante,
            result.empresa,
            result.link,
            "OK" if result.status == LinkStatus.OK else "ERRO",
        ]

    @staticmethod
    def _technical_row(result: ValidationResult) -> list[object]:
        return [
            result.participante,
            result.empresa,
            result.link,
            result.status.value,
            result.http_status,
            result.final_url,
            result.response_time_seconds,
            result.rule_name,
            result.evidence,
            result.technical_error,
        ]


class _Sheet:
    def __init__(
        self,
        name: str,
        columns: tuple[str, ...],
        rows: list[list[object]],
        *,
        result_sheet: bool = False,
    ) -> None:
        self.name = name
        self.columns = columns
        self.rows = rows
        self.result_sheet = result_sheet


def _write_csv(path: Path, columns: tuple[str, ...], rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(columns)
        writer.writerows(rows)


def _write_xlsx(path: Path, sheets: list[_Sheet]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    error_fill = PatternFill("solid", fgColor="FFC7CE")

    for sheet in sheets:
        worksheet = workbook.create_sheet(sheet.name)
        worksheet.freeze_panes = "A2"
        worksheet.append(list(sheet.columns))
        for row in sheet.rows:
            worksheet.append(row)

        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        if sheet.result_sheet:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=4)
                cell.fill = ok_fill if cell.value == "OK" else error_fill

        for index, width in enumerate(_column_widths(sheet), 1):
            worksheet.column_dimensions[_column_name(index)].width = width

    workbook.save(path)


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _column_widths(sheet: _Sheet) -> list[int]:
    widths = (
        [28, 24, 50, 12] if sheet.result_sheet else [28, 24, 42, 10, 18, 16, 28, 12, 36, 14, 28]
    )
    return widths[: len(sheet.columns)]


def _details_row(result: ValidationResult) -> list[object]:
    return [
        result.participante,
        result.empresa,
        result.link,
        "OK" if result.status == LinkStatus.OK else "ERRO",
        result.status.value,
        result.rule_name or "",
        result.evidence or "",
        result.http_status if result.http_status is not None else "",
        result.final_url or "",
        round(result.response_time_seconds * 1000)
        if result.response_time_seconds is not None
        else "",
        result.technical_error or "",
    ]


def _ensure_can_write(path: Path, overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(
            f"Arquivo de saida ja existe: {path}. Use --overwrite para substituir."
        )
