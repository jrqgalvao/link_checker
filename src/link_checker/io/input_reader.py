from __future__ import annotations

import csv
from pathlib import Path
from urllib.parse import urlparse

from link_checker.models import InputLinkRecord

_SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
_REQUIRED_CSV_COLUMNS = ("participante", "email", "empresa", "evento_esperado", "link")


class InputReader:
    def read(self, path: Path) -> list[InputLinkRecord]:
        if not path.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

        suffix = path.suffix.lower()

        if suffix == ".csv":
            return self._read_csv(path)
        if suffix in _SUPPORTED_EXCEL_EXTENSIONS:
            return self._read_excel(path, suffix)
        raise ValueError(f"Extensao nao suportada: {suffix}. Use .csv, .xlsx, .xlsm ou .xls.")

    def _read_csv(self, path: Path) -> list[InputLinkRecord]:
        with path.open(newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            missing = [col for col in _REQUIRED_CSV_COLUMNS if col not in (reader.fieldnames or [])]
            if missing:
                raise ValueError(f"Colunas obrigatorias ausentes: {', '.join(missing)}")
            return [
                InputLinkRecord(
                    participante=str(row.get("participante") or "").strip(),
                    email=str(row.get("email") or "").strip(),
                    empresa=str(row.get("empresa") or "").strip(),
                    evento_esperado=str(row.get("evento_esperado") or "").strip(),
                    link=str(row.get("link") or "").strip(),
                )
                for row in reader
            ]

    def _read_excel(self, path: Path, suffix: str) -> list[InputLinkRecord]:
        try:
            rows = _read_xls_rows(path) if suffix == ".xls" else _read_xlsx_rows(path)
        except Exception as exc:
            raise ValueError(f"Nao foi possivel ler a planilha Excel: {exc}") from exc

        if not rows:
            raise ValueError("Planilha Excel vazia na primeira aba.")

        width = max(len(row) for row in rows)
        if width < 3:
            raise ValueError(
                "Planilha Excel deve ter pelo menos 3 colunas: "
                "nome do participante, nome da empresa e link."
            )

        first_row = [_cell(rows[0], i).lower() for i in range(3)]
        if first_row[0].startswith("nome") and first_row[2] == "link":
            raise ValueError(
                "A planilha Excel nao deve ter cabecalho. Remova a primeira linha de cabecalho."
            )

        if self._looks_like_title_row(rows):
            raise ValueError(
                "A primeira linha parece titulo ou observacao, nao um registro operacional."
            )

        records: list[InputLinkRecord] = []
        for row in rows:
            participante = _cell(row, 0)
            empresa = _cell(row, 1)
            link = _cell(row, 2)

            if not participante and not empresa and not link:
                continue

            records.append(
                InputLinkRecord(
                    participante=participante,
                    email="",
                    empresa=empresa,
                    evento_esperado="",
                    link=link,
                )
            )

        if not records:
            raise ValueError("Planilha Excel vazia.")

        return records

    @staticmethod
    def _looks_like_title_row(rows: list[list[object]]) -> bool:
        if len(rows) < 2:
            return False
        first_link = _cell(rows[0], 2)
        second_link = _cell(rows[1], 2)
        return not _has_http_scheme(first_link) and _has_http_scheme(second_link)


def _read_xlsx_rows(path: Path) -> list[list[object]]:
    from openpyxl import load_workbook

    value_workbook = load_workbook(path, read_only=False, data_only=True)
    formula_workbook = load_workbook(path, read_only=False, data_only=False)
    value_sheet = value_workbook.worksheets[0]
    formula_sheet = formula_workbook.worksheets[0]
    rows: list[list[object]] = []
    for value_cells, formula_cells in zip(
        value_sheet.iter_rows(), formula_sheet.iter_rows(), strict=True
    ):
        row = []
        for value_cell, formula_cell in zip(value_cells, formula_cells, strict=True):
            if value_cell.column == 3:
                row.append(_link_value(value_cell.value, formula_cell))
            else:
                row.append(value_cell.value)
        rows.append(row)
    return rows


def _read_xls_rows(path: Path) -> list[list[object]]:
    import xlrd

    sheet = xlrd.open_workbook(path, formatting_info=True).sheet_by_index(0)
    rows = [sheet.row_values(index) for index in range(sheet.nrows)]
    for (row_index, col_index), hyperlink in getattr(sheet, "hyperlink_map", {}).items():
        if col_index == 2 and row_index < len(rows) and len(rows[row_index]) > col_index:
            url = getattr(hyperlink, "url_or_path", "")
            if url:
                rows[row_index][col_index] = url
    return rows


def _cell(row: list[object], index: int) -> str:
    if index >= len(row):
        return ""
    return str(row[index] or "").strip()


def _link_value(value: object, formula_cell) -> object:
    if formula_cell.hyperlink and formula_cell.hyperlink.target:
        return formula_cell.hyperlink.target
    if isinstance(formula_cell.value, str):
        formula_link = _hyperlink_formula_target(formula_cell.value)
        if formula_link:
            return formula_link
    return value


def _hyperlink_formula_target(value: str) -> str:
    formula = value.strip()
    if not formula[:11].lower().startswith(("=hyperlink(", "=hiperlink(")):
        return ""

    start = formula.find('"')
    if start < 0:
        return ""

    chars: list[str] = []
    index = start + 1
    while index < len(formula):
        char = formula[index]
        if char == '"':
            if index + 1 < len(formula) and formula[index + 1] == '"':
                chars.append('"')
                index += 2
                continue
            break
        chars.append(char)
        index += 1
    return "".join(chars).strip()


def _has_http_scheme(value: str) -> bool:
    return urlparse(value).scheme in {"http", "https"}
