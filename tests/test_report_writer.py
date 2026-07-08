from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from link_checker.enums import LinkStatus
from link_checker.io.report_writer import ReportWriter
from link_checker.models import ValidationResult


def _result(status: LinkStatus) -> ValidationResult:
    return ValidationResult(
        participante="Ana",
        email="",
        empresa="ACME",
        evento_esperado="",
        link="https://x.test",
        status=status,
        http_status=404 if status == LinkStatus.MORTO_404 else 200,
        final_url="https://final.test",
        response_time_seconds=1.25,
        rule_name="test_rule",
        evidence="evidencia",
        technical_error="erro tecnico" if status == LinkStatus.ERRO_TECNICO else None,
    )


@pytest.fixture
def writer() -> ReportWriter:
    return ReportWriter()


@pytest.fixture
def all_status_results() -> list[ValidationResult]:
    return [_result(s) for s in LinkStatus]


def _read_xlsx(path: Path) -> list[dict[str, object]]:
    ws = load_workbook(path, data_only=True)["resultado"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value) for value in rows[0]]
    return [dict(zip(headers, row, strict=True)) for row in rows[1:]]


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as file:
        return list(csv.DictReader(file))


class TestXlsxSimples:
    def test_gera_xlsx(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_contem_quatro_colunas(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        rows = _read_xlsx(path)
        assert list(rows[0]) == [
            "NOME DO PARTICIPANTE",
            "NOME DA EMPRESA",
            "LINK",
            "RESULTADO",
        ]

    def test_uma_linha_por_resultado(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        rows = _read_xlsx(path)
        assert len(rows) == len(LinkStatus)

    def test_ok_vira_ok(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.OK)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "OK"

    def test_invalido_suporte_vira_erro(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.INVALIDO_SUPORTE)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "ERRO"

    def test_morto_404_vira_erro(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.MORTO_404)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "ERRO"

    def test_erro_http_vira_erro(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.ERRO_HTTP)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "ERRO"

    def test_timeout_vira_erro(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.TIMEOUT)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "ERRO"

    def test_erro_tecnico_vira_erro(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.ERRO_TECNICO)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "ERRO"

    def test_indeterminado_vira_erro(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.INDETERMINADO)], path)
        rows = _read_xlsx(path)
        assert rows[0]["RESULTADO"] == "ERRO"

    def test_colunas_tecnicas_nao_aparecem(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        rows = _read_xlsx(path)
        for col in rows[0]:
            assert col not in (
                "STATUS",
                "REGRA",
                "EVIDENCIA",
                "STATUS_HTTP",
                "URL_FINAL",
                "TEMPO_RESPOSTA_MS",
                "ERRO_TECNICO",
            )

    def test_cria_diretorio_saida(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "subdir" / "report.xlsx"
        writer.write(all_status_results, path)
        assert path.exists()

    def test_nao_sobrescreve_saida_existente_por_padrao(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        path.write_text("existente")

        with pytest.raises(FileExistsError, match="Arquivo de saida ja existe"):
            writer.write(all_status_results, path)

    def test_sobrescreve_saida_existente_quando_solicitado(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        path.write_text("existente")

        writer.write(all_status_results, path, overwrite=True)

        assert path.exists()

    def test_tem_aba_resultado(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        assert "resultado" in wb.sheetnames

    def test_tem_aba_detalhes(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        assert "detalhes" in wb.sheetnames

    def test_resultado_e_primeira_aba(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        assert wb.sheetnames[0] == "resultado"

    def test_detalhes_colunas_tecnicas(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        ws = wb["detalhes"]
        cols = [cell.value for cell in ws[1]]
        assert cols == [
            "NOME DO PARTICIPANTE",
            "NOME DA EMPRESA",
            "LINK",
            "RESULTADO",
            "STATUS_INTERNO",
            "REGRA",
            "EVIDENCIA",
            "STATUS_HTTP",
            "URL_FINAL",
            "TEMPO_RESPOSTA_MS",
            "ERRO_TECNICO",
        ]

    def test_detalhes_contem_status_interno(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        ws = wb["detalhes"]
        statuses = {ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)}
        assert "OK" in statuses
        assert "MORTO_404" in statuses
        assert "ERRO_TECNICO" in statuses

    def test_detalhes_evidencia(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        ws = wb["detalhes"]
        values = {ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)}
        assert "evidencia" in values

    def test_detalhes_erro_tecnico(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        ws = wb["detalhes"]
        values = {ws.cell(row=r, column=11).value for r in range(2, ws.max_row + 1)}
        assert "erro tecnico" in values

    def test_cabecalho_com_negrito(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsx"
        writer.write(all_status_results, path)
        wb = load_workbook(path)
        ws = wb.active
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_ok_recebe_fill_verde(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.OK)], path)
        wb = load_workbook(path)
        ws = wb.active
        cell = ws.cell(row=2, column=4)
        assert cell.value == "OK"
        assert cell.fill.start_color.rgb.endswith("C6EFCE")

    def test_erro_recebe_fill_vermelho(self, writer: ReportWriter, tmp_path: Path) -> None:
        path = tmp_path / "report.xlsx"
        writer.write([_result(LinkStatus.MORTO_404)], path)
        wb = load_workbook(path)
        ws = wb.active
        cell = ws.cell(row=2, column=4)
        assert cell.value == "ERRO"
        assert cell.fill.start_color.rgb.endswith("FFC7CE")


class TestCsvSimples:
    def test_csv_legado_mesmas_colunas(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.csv"
        writer.write(all_status_results, path)
        rows = _read_csv(path)
        assert list(rows[0]) == [
            "NOME DO PARTICIPANTE",
            "NOME DA EMPRESA",
            "LINK",
            "RESULTADO",
        ]


class TestRelatorioTecnico:
    def test_gera_relatorio_tecnico_com_campos_de_debug(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "technical.csv"
        writer.write_technical(all_status_results, path)

        rows = _read_csv(path)

        assert list(rows[0]) == [
            "participante",
            "empresa",
            "link",
            "status",
            "http_status",
            "final_url",
            "response_time_seconds",
            "rule_name",
            "evidence",
            "technical_error",
        ]
        assert "MORTO_404" in {row["status"] for row in rows}

    def test_relatorio_tecnico_respeita_protecao_de_sobrescrita(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "technical.csv"
        path.write_text("existente")

        with pytest.raises(FileExistsError, match="Arquivo de saida ja existe"):
            writer.write_technical(all_status_results, path)


class TestErrors:
    def test_rejeita_extensao_nao_suportada(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.txt"
        with pytest.raises(ValueError, match="Extensao nao suportada"):
            writer.write(all_status_results, path)

    def test_rejeita_xlsm_para_evitar_confusao_com_macros(
        self, writer: ReportWriter, all_status_results: list[ValidationResult], tmp_path: Path
    ) -> None:
        path = tmp_path / "report.xlsm"
        with pytest.raises(ValueError, match="Use .csv ou .xlsx"):
            writer.write(all_status_results, path)
