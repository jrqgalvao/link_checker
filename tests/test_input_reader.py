from __future__ import annotations

from pathlib import Path
from unittest import mock

import pytest
from openpyxl import Workbook

from link_checker.io.input_reader import InputReader


@pytest.fixture
def reader() -> InputReader:
    return InputReader()


def _create_excel(tmp_path: Path, filename: str, rows: list[list[str]]) -> Path:
    path = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


class TestExcelSemCabecalho:
    def test_le_xlsx_sem_cabecalho(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "test.xlsx",
            [
                ["Ana", "Empresa A", "https://example.com/link1"],
                ["Beto", "Empresa B", "https://example.com/link2"],
            ],
        )
        records = reader.read(path)
        assert len(records) == 2
        assert records[0].participante == "Ana"
        assert records[0].empresa == "Empresa A"
        assert records[0].link == "https://example.com/link1"

    def test_le_xlsm_sem_cabecalho(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "test.xlsm",
            [
                ["Carlos", "Empresa C", "https://example.com/link3"],
            ],
        )
        records = reader.read(path)
        assert len(records) == 1
        assert records[0].participante == "Carlos"

    def test_coluna_a_mapeia_participante(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "map.xlsx",
            [
                ["Joao", "Foo Ltda", "https://x.test/1"],
            ],
        )
        records = reader.read(path)
        assert records[0].participante == "Joao"

    def test_coluna_b_mapeia_empresa(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "map.xlsx",
            [
                ["Joao", "Foo Ltda", "https://x.test/1"],
            ],
        )
        records = reader.read(path)
        assert records[0].empresa == "Foo Ltda"

    def test_coluna_c_mapeia_link(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "map.xlsx",
            [
                ["Joao", "Foo Ltda", "https://x.test/1"],
            ],
        )
        records = reader.read(path)
        assert records[0].link == "https://x.test/1"

    def test_coluna_c_usa_hyperlink_quando_texto_nao_e_url(
        self, reader: InputReader, tmp_path: Path
    ) -> None:
        path = tmp_path / "hyperlink.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Ana", "Empresa A", "Click here"])
        ws["C1"].hyperlink = "https://example.com/real-link"
        wb.save(path)

        records = reader.read(path)

        assert records[0].link == "https://example.com/real-link"

    def test_coluna_c_usa_formula_hyperlink_quando_texto_nao_e_url(
        self, reader: InputReader, tmp_path: Path
    ) -> None:
        path = tmp_path / "formula_hyperlink.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["Ana", "Empresa A", '=HYPERLINK("https://example.com/formula", "Click here")'])
        wb.save(path)

        records = reader.read(path)

        assert records[0].link == "https://example.com/formula"

    def test_ignora_colunas_extras(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "extra.xlsx",
            [
                ["Ana", "Empresa A", "https://example.com/link", "extra1", "extra2"],
            ],
        )
        records = reader.read(path)
        assert len(records) == 1
        assert records[0].participante == "Ana"

    def test_ignora_linhas_totalmente_vazias(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "empty.xlsx",
            [
                ["Ana", "Empresa A", "https://example.com/link1"],
                ["", "", ""],
                ["", "", ""],
                ["Beto", "Empresa B", "https://example.com/link2"],
            ],
        )
        records = reader.read(path)
        assert len(records) == 2
        assert records[0].participante == "Ana"
        assert records[1].participante == "Beto"

    def test_preserva_linhas_parcialmente_vazias(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "partial.xlsx",
            [
                ["Ana", "", "https://example.com/link"],
                ["", "Empresa B", "https://example.com/link2"],
            ],
        )
        records = reader.read(path)
        assert len(records) == 2
        assert records[0].empresa == ""
        assert records[1].participante == ""

    def test_remove_espacos_extras(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "strip.xlsx",
            [
                ["  Ana  ", "  Empresa A  ", "  https://example.com/link  "],
            ],
        )
        records = reader.read(path)
        assert records[0].participante == "Ana"
        assert records[0].empresa == "Empresa A"
        assert records[0].link == "https://example.com/link"

    def test_menos_de_tres_colunas_falha(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "few.xlsx",
            [
                ["Ana", "Empresa A"],
            ],
        )
        with pytest.raises(ValueError, match="pelo menos 3 colunas"):
            reader.read(path)

    def test_planilha_vazia_falha(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(tmp_path, "empty.xlsx", [])
        with pytest.raises(ValueError, match="Planilha Excel vazia"):
            reader.read(path)

    def test_cabecalho_antigo_rejeitado(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "header.xlsx",
            [
                ["NOME DO PARTICIPANTE", "NOME DA EMPRESA", "LINK"],
                ["Ana", "Empresa A", "https://example.com/link"],
            ],
        )
        with pytest.raises(ValueError, match="nao deve ter cabecalho"):
            reader.read(path)

    def test_nao_valida_url(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "nourl.xlsx",
            [
                ["Ana", "Empresa A", "not-a-valid-url"],
            ],
        )
        records = reader.read(path)
        assert records[0].link == "not-a-valid-url"

    def test_linha_com_cabecalho_parcial_rejeitada(
        self, reader: InputReader, tmp_path: Path
    ) -> None:
        path = _create_excel(
            tmp_path,
            "partial_header.xlsx",
            [
                ["NOME DO", "NOME DA", "LINK"],
                ["Ana", "Empresa A", "https://example.com/link"],
            ],
        )
        with pytest.raises(ValueError, match="nao deve ter cabecalho"):
            reader.read(path)

    def test_primeira_linha_titulo_rejeitada(self, reader: InputReader, tmp_path: Path) -> None:
        path = _create_excel(
            tmp_path,
            "title.xlsx",
            [
                ["Relatorio de links", "", ""],
                ["Ana", "Empresa A", "https://example.com/link"],
            ],
        )
        with pytest.raises(ValueError, match="primeira linha"):
            reader.read(path)

    def test_primeira_aba_vazia_rejeitada_mesmo_com_dados_na_segunda(
        self, reader: InputReader, tmp_path: Path
    ) -> None:
        path = tmp_path / "second_sheet.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Vazia"
        second = wb.create_sheet("Dados")
        second.append(["Ana", "Empresa A", "https://example.com/link"])
        wb.save(path)

        with pytest.raises(ValueError, match="primeira aba"):
            reader.read(path)


class TestLegadoCsv:
    def test_csv_continua_funcionando(self, reader: InputReader, tmp_path: Path) -> None:
        path = tmp_path / "test.csv"
        path.write_text(
            "participante,email,empresa,evento_esperado,link\nAna,,A,,https://x.test\n",
            encoding="utf-8",
        )
        records = reader.read(path)
        assert len(records) == 1
        assert records[0].participante == "Ana"

    def test_rejeita_extensao_nao_suportada(self, reader: InputReader, tmp_path: Path) -> None:
        path = tmp_path / "test.txt"
        path.write_text("")
        with pytest.raises(ValueError, match="Extensao nao suportada"):
            reader.read(path)

    def test_rejeita_arquivo_inexistente(self, reader: InputReader) -> None:
        path = Path(r"C:\nonexistent_folder\file.xlsx")
        with pytest.raises(FileNotFoundError, match="Arquivo nao encontrado"):
            reader.read(path)

    def test_rejeita_excel_corrompido(self, reader: InputReader, tmp_path: Path) -> None:
        path = tmp_path / "broken.xlsx"
        path.write_bytes(b"not an excel file")

        with pytest.raises(ValueError, match="Nao foi possivel ler"):
            reader.read(path)

    def test_accepts_xls_extension(self, reader: InputReader, tmp_path: Path) -> None:
        path = tmp_path / "legacy.xls"
        path.write_bytes(b"fake xls")

        class FakeSheet:
            nrows = 1

            def row_values(self, _index):
                return ["Ana", "Empresa A", "https://example.com/link"]

        class FakeBook:
            def sheet_by_index(self, index):
                assert index == 0
                return FakeSheet()

        with mock.patch("xlrd.open_workbook", return_value=FakeBook()) as open_workbook:
            records = reader.read(path)

        open_workbook.assert_called_once_with(path, formatting_info=True)
        assert records[0].participante == "Ana"

    def test_xls_coluna_c_usa_hyperlink_quando_texto_nao_e_url(
        self, reader: InputReader, tmp_path: Path
    ) -> None:
        path = tmp_path / "legacy_hyperlink.xls"
        path.write_bytes(b"fake xls")

        class FakeHyperlink:
            url_or_path = "https://example.com/xls-real"

        class FakeSheet:
            nrows = 1
            hyperlink_map = {(0, 2): FakeHyperlink()}

            def row_values(self, _index):
                return ["Ana", "Empresa A", "Click here"]

        class FakeBook:
            def sheet_by_index(self, index):
                assert index == 0
                return FakeSheet()

        with mock.patch("xlrd.open_workbook", return_value=FakeBook()) as open_workbook:
            records = reader.read(path)

        open_workbook.assert_called_once_with(path, formatting_info=True)
        assert records[0].link == "https://example.com/xls-real"
